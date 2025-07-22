#!/usr/bin/env python3
"""
Module for exporting transcript data to Excel format.
"""
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def calculate_column_widths(worksheet) -> None:
    """
    Calculate and set optimal column widths based on content.
    
    Args:
        worksheet: The worksheet to adjust column widths for
    """
    # Set minimum and maximum column widths
    min_width = 8
    max_width = 50
    
    # Iterate through all columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Find the maximum length of content in the column
        for cell in column:
            try:
                # Get the length of the cell value
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Add some padding and set the column width
        adjusted_width = (max_length + 2) * 1.1  # Add padding and a little extra
        
        # Apply min and max width constraints
        if adjusted_width < min_width:
            adjusted_width = min_width
        elif adjusted_width > max_width:
            adjusted_width = max_width
            
        # Set the column width
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_excel(
    data: List[Dict[str, Any]], 
    output_file: str, 
    coach_color: str = 'D8E4F0', 
    client_color: str = 'F0F0F0',  # Light gray for client rows
    font_size: int = 16,           # Slightly smaller default font size
) -> None:
    """
    Generate an Excel file from the parsed data with formatting.
    
    Args:
        data: List of dictionaries with time, speaker, and content
        output_file: Path to save the Excel file
        coach_color: Hex color code to highlight coach rows (default: light blue D8E4F0)
        client_color: Hex color code for client rows (default: light gray F0F0F0)
        font_size: Default font size for all cells (default: 14)
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"
    
    # Add headers
    headers = ['Time', 'Role', 'Content']
    ws.append(headers)
    
    # Add data rows
    for item in data:
        ws.append([item['time'], item['speaker'], item['content']])
    
    # Set initial column widths (will be adjusted later)
    ws.column_dimensions['A'].width = 15  # Time column
    ws.column_dimensions['B'].width = 15  # Role column
    ws.column_dimensions['C'].width = 40  # Content column (initial width)
    
    # Create styles
    header_font = Font(bold=True, size=font_size + 2, color='FFFFFF')  # White text for header
    cell_font = Font(size=font_size)
    alignment = Alignment(wrap_text=True, vertical='top')
    
    # Create fills
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')  # Blue header
    coach_fill = PatternFill(start_color=coach_color, end_color=coach_color, fill_type='solid')
    client_fill = PatternFill(start_color=client_color, end_color=client_color, fill_type='solid')
    
    # Apply styles to header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply styles to data cells
    for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=3):
        speaker = row[1].value  # Column B contains the speaker
        
        for cell in row:
            cell.font = cell_font
            cell.alignment = alignment
            
            # Apply fill based on speaker role
            if speaker == 'Coach':
                cell.fill = coach_fill
            elif speaker == 'Client':
                cell.fill = client_fill
    
    # Auto-adjust column widths based on content
    calculate_column_widths(ws)
    
    # Calculate and set row heights based on content
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Start from row 2 (skip header)
        max_lines = 1
        for cell in row:
            if cell.value:
                # Estimate number of lines based on content length and column width
                col_width = ws.column_dimensions[cell.column_letter].width
                if col_width > 0:  # Make sure column width is valid
                    # Rough estimate: ~10 characters per inch, 1 line per 2 inches
                    content_length = len(str(cell.value))
                    estimated_lines = max(1, int((content_length / (col_width * 1.5)) + 0.5))
                    max_lines = max(max_lines, estimated_lines)
        
        # Set row height (15 points per line, minimum 30 points, maximum 300 points)
        row_height = min(max(30, max_lines * 15), 300)
        ws.row_dimensions[row_idx].height = row_height
    
    # Set header row height
    ws.row_dimensions[1].height = 30  # Fixed height for header
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(output_file)
