#!/usr/bin/env python3
import warnings
warnings.filterwarnings("ignore", category=RuntimeWarning, module="importlib._bootstrap")

import argparse
import re
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def parse_vtt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    pattern = r'(\d{2}:\d{2}:\d{2}\.\d{3}) --> (\d{2}:\d{2}:\d{2}\.\d{3})\n<v ([^>]+)>(.+?)</v>'
    matches = re.findall(pattern, content, re.DOTALL)

    parsed_data = []
    for start_time, _, speaker, text in matches:
        parsed_data.append({
            'time': start_time,
            'speaker': speaker.strip(),
            'content': text.strip().replace('\n', ' ')
        })

    return parsed_data

def consolidate_speakers(data):
    consolidated = []
    current_speaker = None
    current_content = []
    current_time = None

    for item in data:
        if item['speaker'] != current_speaker:
            if current_speaker:
                consolidated.append({
                    'time': current_time,
                    'speaker': current_speaker,
                    'content': ' '.join(current_content)
                })
            current_speaker = item['speaker']
            current_content = [item['content']]
            current_time = item['time']
        else:
            current_content.append(item['content'])

    if current_speaker:
        consolidated.append({
            'time': current_time,
            'speaker': current_speaker,
            'content': ' '.join(current_content)
        })

    return consolidated

def replace_names(data, coach, client):
    for item in data:
        if item['speaker'] == coach:
            item['speaker'] = 'Coach'
        elif item['speaker'] == client:
            item['speaker'] = 'Client'
    return data

def generate_markdown(data, content_width=80):
    """
    Generate a Markdown table from the parsed data.
    
    Args:
        data: List of dictionaries with time, speaker, and content
        content_width: Maximum width for content column before wrapping (default: 80)
    
    Returns:
        Markdown table as a string
    """
    # Create header
    markdown = "| Time | Role | Content |\n"
    markdown += "| ---- | ---- | ------- |\n"
    
    # Process each row
    for item in data:
        # Get the content and wrap it if needed
        content = item['content']
        
        # Format the content to fit within content_width
        if len(content) > content_width:
            # More natural word-based wrapping instead of character-based
            words = content.split()
            wrapped_content = []
            current_line = ""
            
            for word in words:
                if len(current_line) + len(word) + 1 <= content_width:
                    if current_line:
                        current_line += " " + word
                    else:
                        current_line = word
                else:
                    wrapped_content.append(current_line)
                    current_line = word
            
            if current_line:
                wrapped_content.append(current_line)
                
            content = "<br>".join(wrapped_content)
        
        # Add row to markdown table with special formatting for Coach
        if item['speaker'] == 'Coach':
            markdown += f"| {item['time']} | **{item['speaker']}** | {content} |\n"
        else:
            markdown += f"| {item['time']} | {item['speaker']} | {content} |\n"
    
    return markdown

def generate_excel(data, output_file, coach_color='D8E4F0', font_size=16, content_width=40):
    """
    Generate an Excel file from the parsed data with formatting.
    
    Args:
        data: List of dictionaries with time, speaker, and content
        output_file: Path to save the Excel file
        coach_color: Hex color code to highlight coach rows (default: light blue D8E4F0)
        font_size: Default font size for all cells (default: 16)
        content_width: Width of the content column in characters (default: 40)
    """
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel writer with openpyxl engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name='Transcript')
        
        # Get the worksheet
        worksheet = writer.sheets['Transcript']
        
        # Set default font size for the entire worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = openpyxl.styles.Font(size=font_size)
        
        # Get column indices
        time_col = 1  # Excel columns are 1-indexed
        speaker_col = 2
        content_col = 3
        
        # Set column widths - Excel column width units are based on the width of '0' character in the default font
        # For content column, use the specified content_width directly
        worksheet.column_dimensions[get_column_letter(content_col)].width = content_width
        
        # Set reasonable widths for time and speaker columns
        worksheet.column_dimensions[get_column_letter(time_col)].width = 15
        worksheet.column_dimensions[get_column_letter(speaker_col)].width = 12
        
        # Auto-adjust row heights to fit content
        for row_idx in range(1, len(data) + 2):
            # Get the maximum number of lines in any cell in this row
            max_lines = 1
            for col_idx in range(1, 4):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_text = str(cell.value)
                    # For content column, calculate lines more accurately
                    if col_idx == content_col and len(cell_text) > 0:
                        # Calculate average characters per line based on font size
                        # Smaller fonts can fit more characters per line width
                        chars_per_line = content_width * (12 / font_size) * 1.8
                        
                        # Count approximate number of lines needed
                        # Consider both width and character count
                        if len(cell_text) > chars_per_line:
                            # Count newlines in text and add line breaks for long content
                            newline_count = cell_text.count('\n') + 1
                            wrapped_lines = len(cell_text) / chars_per_line
                            num_lines = max(newline_count, wrapped_lines)
                            max_lines = max(max_lines, int(num_lines) + 1)  # Add a buffer line
            
            # Set row height based on font size and number of lines
            # Adjust multiplier based on font size to ensure all content is visible
            height_multiplier = 2.2 if font_size <= 14 else 2.5
            row_height = max_lines * (font_size * height_multiplier)
            worksheet.row_dimensions[row_idx].height = row_height
        
        # Enable text wrapping for content column
        for row in range(2, len(data) + 2):  # +2 because Excel is 1-indexed and we have a header row
            cell = worksheet.cell(row=row, column=content_col)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # Apply color to rows where speaker is "Coach"
        coach_fill = openpyxl.styles.PatternFill(
            start_color=coach_color,
            end_color=coach_color,
            fill_type='solid'
        )
        
        # Apply formatting to data rows (skip header)
        for row_idx, row_data in enumerate(data, start=2):  # Start from row 2 (after header)
            if row_data['speaker'] == 'Coach':
                # Apply fill to all cells in the row
                for col in range(1, 4):  # Columns 1, 2, 3
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.fill = coach_fill

def main():
    parser = argparse.ArgumentParser(description='Convert VTT to Markdown or Excel')
    parser.add_argument('input_file', help='Input VTT file')
    parser.add_argument('output_file', help='Output file (use .md for Markdown or .xlsx for Excel)')
    parser.add_argument('-Coach', help='Name of the coach')
    parser.add_argument('-Client', help='Name of the client')
    parser.add_argument('-color', default='D8E4F0', help='Hex color code for highlighting coach rows (default: D8E4F0 light blue)')
    parser.add_argument('-font_size', type=int, default=16, help='Font size for Excel output (default: 16)')
    parser.add_argument('-content_width', type=int, default=160, help='Width of content column in characters (default: 160 for Excel, 80 for Markdown)')
    args = parser.parse_args()

    data = parse_vtt(args.input_file)
    consolidated_data = consolidate_speakers(data)

    if args.Coach and args.Client:
        consolidated_data = replace_names(consolidated_data, args.Coach, args.Client)

    if args.output_file.endswith('.md'):
        # For Markdown, use a more reasonable content width if not explicitly set
        md_content_width = min(args.content_width, 80) if args.content_width != 160 else 80
        output = generate_markdown(consolidated_data, md_content_width)
        with open(args.output_file, 'w', encoding='utf-8') as f:
            f.write(output)
    elif args.output_file.endswith('.xlsx'):
        generate_excel(consolidated_data, args.output_file, args.color, args.font_size, args.content_width)
    else:
        print("Unsupported output format. Use .md for Markdown or .xlsx for Excel.")

if __name__ == "__main__":
    main()
