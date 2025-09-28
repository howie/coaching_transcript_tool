#!/usr/bin/env python3
"""
Module for exporting transcript data to Excel format.
"""

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def calculate_column_widths(worksheet) -> None:
    """
    Calculate and set optimal column widths based on content.

    Args:
        worksheet: The worksheet to adjust column widths for
    """
    # Set minimum and maximum column widths
    min_width = 8
    max_width = 50

    # Iterate through all columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        # Find the maximum length of content in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Add some padding and set the column width
        adjusted_width = (max_length + 2) * 1.1  # Add padding and a little extra

        # Apply min and max width constraints
        if adjusted_width < min_width:
            adjusted_width = min_width
        elif adjusted_width > max_width:
            adjusted_width = max_width

        # Set the column width
        worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_excel(
    data: List[Dict[str, Any]],
    coach_color: str = "D8E4F0",
    client_color: str = "F0F0F0",  # Light gray for client rows
    font_size: int = 16,  # Slightly smaller default font size
) -> io.BytesIO:
    """
    Generate an Excel file from the parsed data with formatting.

    Args:
        data: List of dictionaries with time, speaker, and content.
        coach_color: Hex color code to highlight coach rows (default: light blue D8E4F0).
        client_color: Hex color code for client rows (default: light gray F0F0F0).
        font_size: Default font size for all cells (default: 14).

    Returns:
        An in-memory BytesIO buffer containing the Excel file.
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"

    # Add headers
    headers = ["時間", "角色", "內容"]
    ws.append(headers)

    # Add data rows
    for item in data:
        ws.append([item["time"], item["speaker"], item["content"]])

    # Set initial column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40

    # Apply styles and formatting
    _apply_styles(ws, data, font_size, coach_color, client_color)

    # Adjust column widths based on content
    calculate_column_widths(ws)

    # Calculate and set row heights based on content
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2), start=2
    ):  # Start from row 2 (skip header)
        max_lines = 1
        for cell in row:
            if cell.value:
                # Estimate number of lines based on content length and column
                # width
                col_width = ws.column_dimensions[cell.column_letter].width
                if col_width > 0:  # Make sure column width is valid
                    # Rough estimate: ~10 characters per inch, 1 line per 2
                    # inches
                    content_length = len(str(cell.value))
                    estimated_lines = max(
                        1, int((content_length / (col_width * 1.5)) + 0.5)
                    )
                    max_lines = max(max_lines, estimated_lines)

        # Set row height (15 points per line, minimum 30 points, maximum 300
        # points)
        row_height = min(max(30, max_lines * 15), 300)
        ws.row_dimensions[row_idx].height = row_height

    # Set header row height
    ws.row_dimensions[1].height = 30  # Fixed height for header

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save the workbook to an in-memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _apply_styles(ws, data, font_size, coach_color, client_color):
    """Helper function to apply styles to the worksheet."""
    header_font = Font(bold=True, size=font_size + 2, color="FFFFFF")
    cell_font = Font(size=font_size)
    alignment = Alignment(wrap_text=True, vertical="top")

    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    coach_fill = PatternFill(
        start_color=coach_color, end_color=coach_color, fill_type="solid"
    )
    client_fill = PatternFill(
        start_color=client_color, end_color=client_color, fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row_idx, item in enumerate(data, start=2):
        fill = coach_fill if item["speaker"] in ["Coach", "教練"] else client_fill
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = alignment
