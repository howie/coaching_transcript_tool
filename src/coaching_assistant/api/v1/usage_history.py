"""
Usage History Analytics API endpoints.

Provides historical usage data, trends, predictions, and insights for users.
"""

import logging
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel

from ...core.models.user import User
from ...core.services.usage_tracking_use_case import (
    CreateUsageSnapshotUseCase,
    ExportUsageDataUseCase,
    GetUsageHistoryUseCase,
    GetUsageInsightsUseCase,
    GetUsagePredictionsUseCase,
    GetUsageTrendsUseCase,
)
from .auth import get_current_user_dependency
from .dependencies import (
    get_export_usage_data_use_case,
    get_usage_history_use_case,
    get_usage_insights_use_case,
    get_usage_predictions_use_case,
    get_usage_snapshot_use_case,
    get_usage_trends_use_case,
)

logger = logging.getLogger(__name__)

router = APIRouter()


class UsageHistoryResponse(BaseModel):
    """Response model for usage history data."""

    id: str
    recorded_at: str
    period_type: str
    period_start: str
    period_end: str
    usage_metrics: Dict[str, Any]
    plan_context: Dict[str, Any]
    cost_metrics: Dict[str, Any]
    provider_breakdown: Dict[str, Any]
    export_activity: Dict[str, Any]
    performance_metrics: Dict[str, Any]
    created_at: str
    updated_at: str


class UsageTrendResponse(BaseModel):
    """Response model for usage trend data."""

    date: str
    sessions: int
    minutes: float
    hours: float
    transcriptions: int
    exports: int
    cost: float
    utilization: float
    success_rate: float
    avg_duration: float


class UsagePredictionResponse(BaseModel):
    """Response model for usage predictions."""

    predicted_sessions: int
    predicted_minutes: int
    estimated_limit_date: Optional[str]
    confidence: float
    recommendation: str
    growth_rate: float
    current_trend: str


class UsageInsightResponse(BaseModel):
    """Response model for usage insights."""

    type: str  # "pattern", "optimization", "warning", "trend", "cost"
    title: str
    message: str
    suggestion: str
    priority: str  # "low", "medium", "high"
    action: Optional[str]  # "upgrade", "downgrade", "review", None


class UsageAnalyticsResponse(BaseModel):
    """Response model for comprehensive usage analytics."""

    current_period: Dict[str, Any]
    trends: List[UsageTrendResponse]
    predictions: UsagePredictionResponse
    insights: List[UsageInsightResponse]
    summary: Dict[str, Any]


@router.get("/")
async def usage_history_root():
    """Usage History API root endpoint."""
    return {"message": "Usage History Analytics API v1.0"}


@router.get("/history", response_model=List[UsageHistoryResponse])
async def get_usage_history(
    period: str = Query("30d", description="Time period: 7d, 30d, 3m, 12m, or custom"),
    from_date: Optional[str] = Query(
        None, description="Start date (YYYY-MM-DD) for custom period"
    ),
    to_date: Optional[str] = Query(
        None, description="End date (YYYY-MM-DD) for custom period"
    ),
    period_type: str = Query(
        "daily", description="Aggregation level: hourly, daily, monthly"
    ),
    current_user: User = Depends(get_current_user_dependency),
    usage_history_use_case: GetUsageHistoryUseCase = Depends(
        get_usage_history_use_case
    ),
):
    """
    Get historical usage data for the current user.

    Returns aggregated usage data for specified time period and granularity.
    """
    try:
        logger.info(
            f"📊 Getting usage history for user {current_user.id}, period: {period}"
        )

        # Determine months to retrieve based on period
        months_mapping = {
            "7d": 1,
            "30d": 1,
            "3m": 3,
            "12m": 12,
            "custom": 12,  # Max for custom periods
        }
        months = months_mapping.get(period, 3)

        # Execute use case
        result = usage_history_use_case.execute(user_id=current_user.id, months=months)

        # Convert to response format
        response_data = []
        for history_item in result.get("usage_history", []):
            # Create response with proper field mapping
            usage_history_dict = {
                "id": str(history_item.get("id", "")),
                "recorded_at": history_item.get("recorded_at", ""),
                "period_type": history_item.get("period_type", period_type),
                "period_start": history_item.get("period_start", ""),
                "period_end": history_item.get("period_end", ""),
                "usage_metrics": history_item.get("usage_metrics", {}),
                "plan_context": history_item.get("plan_context", {}),
                "cost_metrics": history_item.get("cost_metrics", {}),
                "provider_breakdown": history_item.get("provider_breakdown", {}),
                "export_activity": history_item.get("export_activity", {}),
                "performance_metrics": history_item.get("performance_metrics", {}),
                "created_at": history_item.get("created_at", ""),
                "updated_at": history_item.get("updated_at", ""),
            }
            response_data.append(UsageHistoryResponse(**usage_history_dict))

        logger.info(
            f"✅ Retrieved {len(response_data)} usage history records for user {current_user.id}"
        )
        return response_data

    except Exception as e:
        logger.error(
            f"❌ Error getting usage history for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve usage history",
        )


@router.get("/trends", response_model=List[UsageTrendResponse])
async def get_usage_trends(
    period: str = Query("30d", description="Time period: 7d, 30d, 3m, 12m"),
    group_by: str = Query("day", description="Grouping: day, week, month"),
    current_user: User = Depends(get_current_user_dependency),
    usage_trends_use_case: GetUsageTrendsUseCase = Depends(get_usage_trends_use_case),
):
    """
    Get usage trends and patterns over time.

    Returns trend data optimized for visualization and analysis.
    """
    try:
        logger.info(f"📈 Getting usage trends for user {current_user.id}")

        # Execute use case
        trends_data = usage_trends_use_case.execute(
            user_id=current_user.id, period=period, group_by=group_by
        )

        # Convert to response format
        response_data = [UsageTrendResponse(**trend) for trend in trends_data]

        logger.info(
            f"✅ Retrieved {len(response_data)} trend data points for user {current_user.id}"
        )
        return response_data

    except Exception as e:
        logger.error(
            f"❌ Error getting usage trends for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve usage trends",
        )


@router.get("/predictions", response_model=UsagePredictionResponse)
async def get_usage_predictions(
    current_user: User = Depends(get_current_user_dependency),
    usage_predictions_use_case: GetUsagePredictionsUseCase = Depends(
        get_usage_predictions_use_case
    ),
):
    """
    Get usage predictions based on historical patterns.

    Returns predicted usage for the next period and recommendations.
    """
    try:
        logger.info(f"🔮 Getting usage predictions for user {current_user.id}")

        # Execute use case
        predictions = usage_predictions_use_case.execute(user_id=current_user.id)

        response_data = UsagePredictionResponse(**predictions)

        logger.info(f"✅ Generated usage predictions for user {current_user.id}")
        return response_data

    except Exception as e:
        logger.error(
            f"❌ Error getting usage predictions for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate usage predictions",
        )


@router.get("/insights", response_model=List[UsageInsightResponse])
async def get_usage_insights(
    current_user: User = Depends(get_current_user_dependency),
    usage_insights_use_case: GetUsageInsightsUseCase = Depends(
        get_usage_insights_use_case
    ),
):
    """
    Get personalized usage insights and recommendations.

    Returns actionable insights based on usage patterns and behavior.
    """
    try:
        logger.info(f"💡 Getting usage insights for user {current_user.id}")

        # Execute use case
        insights = usage_insights_use_case.execute(user_id=current_user.id)

        # Convert to response format
        response_data = [UsageInsightResponse(**insight) for insight in insights]

        logger.info(
            f"✅ Generated {len(response_data)} insights for user {current_user.id}"
        )
        return response_data

    except Exception as e:
        logger.error(
            f"❌ Error getting usage insights for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate usage insights",
        )


@router.get("/analytics", response_model=UsageAnalyticsResponse)
async def get_comprehensive_analytics(
    period: str = Query("30d", description="Time period for analysis"),
    current_user: User = Depends(get_current_user_dependency),
    usage_trends_use_case: GetUsageTrendsUseCase = Depends(get_usage_trends_use_case),
    usage_predictions_use_case: GetUsagePredictionsUseCase = Depends(
        get_usage_predictions_use_case
    ),
    usage_insights_use_case: GetUsageInsightsUseCase = Depends(
        get_usage_insights_use_case
    ),
):
    """
    Get comprehensive usage analytics including trends, predictions, and insights.

    Returns a complete analytics package for dashboard display.
    """
    try:
        logger.info(f"📊 Getting comprehensive analytics for user {current_user.id}")

        # Get all analytics data using use cases
        trends = usage_trends_use_case.execute(
            user_id=current_user.id, period=period, group_by="day"
        )
        predictions = usage_predictions_use_case.execute(user_id=current_user.id)
        insights = usage_insights_use_case.execute(user_id=current_user.id)

        # Calculate current period summary
        current_period = {
            "total_sessions": (
                sum(t["sessions"] for t in trends[-7:]) if trends else 0
            ),
            "total_minutes": (sum(t["minutes"] for t in trends[-7:]) if trends else 0),
            "total_transcriptions": (
                sum(t["transcriptions"] for t in trends[-7:]) if trends else 0
            ),
            "avg_daily_usage": (
                (sum(t["minutes"] for t in trends[-7:]) / 7) if trends else 0
            ),
            "period_start": (
                trends[-7]["date"]
                if len(trends) >= 7
                else (trends[0]["date"] if trends else None)
            ),
            "period_end": trends[-1]["date"] if trends else None,
        }

        # Calculate summary statistics
        summary = {
            "data_points": len(trends),
            "date_range": (
                f"{trends[0]['date']} to {trends[-1]['date']}" if trends else "No data"
            ),
            "total_cost": sum(t["cost"] for t in trends),
            "avg_utilization": (
                sum(t["utilization"] for t in trends) / len(trends) if trends else 0
            ),
            "peak_usage_day": (
                max(trends, key=lambda x: x["minutes"])["date"] if trends else None
            ),
            "insights_count": len(insights),
            "prediction_confidence": predictions["confidence"],
        }

        response_data = UsageAnalyticsResponse(
            current_period=current_period,
            trends=[UsageTrendResponse(**t) for t in trends],
            predictions=UsagePredictionResponse(**predictions),
            insights=[UsageInsightResponse(**i) for i in insights],
            summary=summary,
        )

        logger.info(f"✅ Generated comprehensive analytics for user {current_user.id}")
        return response_data

    except Exception as e:
        logger.error(
            f"❌ Error getting comprehensive analytics for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve comprehensive analytics",
        )


@router.post("/snapshot")
async def create_usage_snapshot(
    period_type: str = Query(
        "daily", description="Snapshot type: hourly, daily, monthly"
    ),
    current_user: User = Depends(get_current_user_dependency),
    usage_snapshot_use_case: CreateUsageSnapshotUseCase = Depends(
        get_usage_snapshot_use_case
    ),
):
    """
    Create a usage snapshot for the current period.

    Manually trigger usage data aggregation (normally done automatically).
    """
    try:
        logger.info(f"📸 Creating usage snapshot for user {current_user.id}")

        # Execute use case
        snapshot = usage_snapshot_use_case.execute(
            user_id=current_user.id, period_type=period_type
        )

        response_data = UsageHistoryResponse(**snapshot.to_dict())

        logger.info(f"✅ Created usage snapshot for user {current_user.id}")
        return response_data

    except Exception as e:
        logger.error(
            f"❌ Error creating usage snapshot for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create usage snapshot",
        )


@router.get("/export")
async def export_usage_data(
    format: str = Query("json", description="Export format: xlsx, txt, json"),
    period: str = Query("30d", description="Time period to export"),
    current_user: User = Depends(get_current_user_dependency),
    export_usage_data_use_case: ExportUsageDataUseCase = Depends(
        get_export_usage_data_use_case
    ),
):
    """
    Export usage data in specified format.

    Returns downloadable usage data for external analysis.
    """
    try:
        logger.info(
            f"📤 Exporting usage data for user {current_user.id}, format: {format}"
        )

        # Execute use case to get export data
        export_result = export_usage_data_use_case.execute(
            user_id=current_user.id, format=format, period=period
        )

        if format.lower() == "xlsx":
            # Create Excel export for usage data
            def export_usage_to_excel(data):
                """Convert usage trends data to Excel format."""
                import io

                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill

                wb = Workbook()
                ws = wb.active
                ws.title = "Usage History"

                # Headers
                headers = [
                    "Date",
                    "Sessions",
                    "Minutes",
                    "Transcriptions",
                    "Exports",
                    "Cost (USD)",
                    "Storage (MB)",
                ]
                ws.append(headers)

                # Style header
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill

                # Add data rows
                for item in data:
                    ws.append(
                        [
                            item.get("date", "N/A"),
                            item.get("sessions", 0),
                            item.get("minutes", 0.0),
                            item.get("transcriptions", 0),
                            item.get("exports_generated", 0),
                            item.get("cost", 0.0),
                            item.get("storage_used_mb", 0.0),
                        ]
                    )

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError, ValueError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer.getvalue()

            # Get trends data from export result
            trends = export_result.get("data", [])
            excel_data = export_usage_to_excel(trends)
            return {
                "format": "xlsx",
                "period": period,
                "data": excel_data,
                "filename": f"usage_history_{period}_{current_user.id}.xlsx",
            }
        elif format.lower() in ["txt", "json"]:
            # Return the data from use case
            return export_result
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported export format. Use xlsx, txt, or json.",
            )

    except Exception as e:
        logger.error(
            f"❌ Error exporting usage data for user {current_user.id}: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export usage data",
        )
